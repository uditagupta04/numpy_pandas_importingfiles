import openpyxl
path = "excel data.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active
cell_obj = sheet_obj.cell(row = 1,column = 3)
print(cell_obj.value)
cell_obj = sheet_obj.cell(row = 1, column = 4)
print(cell_obj.value)