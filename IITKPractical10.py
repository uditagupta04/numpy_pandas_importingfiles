import openpyxl
path = "excel data.xlsx"
wb_obj = openpyxl.load_workbook(path)
sheet_obj = wb_obj.active

row = sheet_obj.max_row
column = sheet_obj.max_column
print("total rows:",row)
print("total columns:",column)
print("\nvalue of first column")
for r in range(1,row+1):
        cell_obj=sheet_obj.cell(row=r,column=1)
        print(cell_obj.value)
for c in range(1,column+1):
        cell_obj=sheet_obj.cell(row=2,column=c)
        print(cell_obj.value,end=" ")

print("\n\n")
for r in range(1,row+1):
        for c in range(1,column+1):
                cell_obj=sheet_obj.cell(row=r,column=c)
                print("%-20s"%cell_obj.value,end="")
        print()